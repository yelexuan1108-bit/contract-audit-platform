"""
AI 股票成交单审核平台
FastAPI Web 应用 — 上传 PDF -> 解析 -> AI 审核 -> 展示报告
"""
import os
import json
import shutil
import asyncio
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, File, UploadFile, Request, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from typing import Optional

from modules.pdf_parser import PDFParser, ContractData
from modules.audit_engine import AuditEngine, AuditReport
from modules.va_checker import VAChecker
from modules.contract_auditor import (
    ContractAuditor,
    STOCK_EN_CLAUSES, STOCK_ZH_HK_CLAUSES, STOCK_ZH_CN_CLAUSES,
    FUND_EN_CLAUSES, FUND_ZH_HK_CLAUSES, FUND_ZH_CN_CLAUSES,
    VA_EN_CLAUSES, VA_ZH_HK_CLAUSES, VA_ZH_CN_CLAUSES,
)

# ===== 应用初始化 =====
app = FastAPI(
    title="Contract Note Audit Platform",
    description="AI 驱动的股票成交单审核平台",
    version="1.0.0",
)

# 静态文件 & 模板
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# 目录
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
RECORDS_DIR = Path("records")
RECORDS_DIR.mkdir(exist_ok=True)

STOCK_SUBTYPES = [
    "order", "orderHk", "companyAction", "companyActionHk",
    "business", "businessHk",
]
VA_SUBTYPES = ["buy", "sale"]
STATEMENT_SUBTYPES = ["monthly", "bbInvest", "southBound"]


# ===== 辅助函数 =====
def get_audit_engine() -> AuditEngine:
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    return AuditEngine(api_key=api_key)

def save_record(data: dict):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    record_path = RECORDS_DIR / f"{timestamp}.json"
    with open(record_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ===== 路由 =====

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse(request, "landing.html", {
        "title": "AI 成交单审核平台",
    })


@app.get("/app", response_class=HTMLResponse)
async def app_page(request: Request):
    return templates.TemplateResponse(request, "index.html", {
        "title": "AI 成交单审核平台",
    })


@app.get("/api/records", response_class=JSONResponse)
async def get_records():
    records = []
    for f in sorted(RECORDS_DIR.glob("*.json"), reverse=True)[:50]:
        try:
            with open(f, "r", encoding="utf-8") as fp:
                data = json.load(fp)
                records.append({
                    "id": f.stem,
                    "file_name": data.get("file_name", ""),
                    "audit_time": data.get("audit_time", ""),
                    "total_findings": data.get("total_findings", 0),
                    "critical_count": data.get("critical_count", 0),
                    "warning_count": data.get("warning_count", 0),
                })
        except Exception:
            continue
    return {"records": records}


@app.get("/api/records/{record_id}", response_class=JSONResponse)
async def get_record(record_id: str):
    record_path = RECORDS_DIR / f"{record_id}.json"
    if not record_path.exists():
        return JSONResponse({"error": "记录不存在"}, status_code=404)
    with open(record_path, "r", encoding="utf-8") as f:
        return json.load(f)


@app.post("/api/chat", response_class=JSONResponse)
async def chat(request: Request):
    body = await request.json()
    question = body.get("question", "")
    contract_data = body.get("contract_data", {})
    api_key = os.environ.get("OPENROUTER_API_KEY", "")

    if not api_key:
        return {"answer": "未配置 API Key，无法使用聊天功能。"}

    import httpx

    has_data = bool(contract_data and any(v for v in contract_data.values() if v))

    if has_data:
        context = json.dumps(contract_data, ensure_ascii=False)
        prompt = f"以下是一份股票成交单的数据：\n{context}\n\n用户问题：{question}\n\n请用简洁专业的语言回答，使用简体中文。"
    else:
        prompt = f"{question}\n\n请用简洁专业的语言回答，使用简体中文。"

    try:
        response = httpx.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "nvidia/nemotron-3-nano-30b-a3b:free",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 512,
            },
            timeout=30,
        )
        answer = response.json()["choices"][0]["message"]["content"]
        return {"answer": answer}
    except Exception as e:
        return {"answer": f"AI 回答失败: {str(e)}"}


@app.post("/api/batch-audit", response_class=JSONResponse)
async def batch_audit_contracts(files: list[UploadFile] = File(...)):
    """批量审核成交单重要提示条款"""
    if len(files) > 200:
        return JSONResponse(
            {"error": f"单次最多上传 200 份文件，当前 {len(files)} 份"},
            status_code=400
        )

    auditor = ContractAuditor()
    results = []
    errors = []

    for file in files:
        if not file.filename or not file.filename.lower().endswith('.pdf'):
            errors.append({"file": file.filename, "error": "非 PDF 文件"})
            continue

        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        safe_name = Path(file.filename).name or "file.pdf"
        file_path = UPLOAD_DIR / f"{ts}_{safe_name}"
        try:
            with open(file_path, "wb") as f:
                f.write(await file.read())
            result = auditor.audit(str(file_path), safe_name)
            data = result.model_dump()
            save_record({"type": "clause_audit", **data})
            results.append(data)
        except Exception as e:
            errors.append({"file": file.filename, "error": str(e)})

    return {
        "success": True,
        "total": len(results),
        "errors": errors,
        "results": results,
    }


@app.get("/api/export/clause-audit", response_class=StreamingResponse)
async def export_clause_audit():
    """导出条款审核结果为 Excel"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()

    # ===== 颜色定义 =====
    GREEN  = PatternFill("solid", fgColor="DCFCE7")
    RED    = PatternFill("solid", fgColor="FEE2E2")
    HEADER = PatternFill("solid", fgColor="030516")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD   = Font(bold=True, size=10)
    NORMAL = Font(size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="E5E7EB")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill   = HEADER
            cell.font   = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_row(ws, row_num, fill=None):
        for cell in ws[row_num]:
            if fill: cell.fill = fill
            cell.font = NORMAL
            cell.alignment = LEFT
            cell.border = BORDER

    # ===== Sheet 1: 汇总 =====
    ws1 = wb.active
    ws1.title = "审核汇总"
    headers1 = [
        "文件名", "审核时间", "文件类型", "市场", "语言", "交易类型",
        "客户姓名", "账户号码", "成交日期",
        "英文条款总数", "英文通过", "英文失败",
        "中文条款总数", "中文通过", "中文失败",
        "总体结论", "失败条款"
    ]
    ws1.append(headers1)
    style_header(ws1)

    # ===== Sheet 2: 英文条款详情 =====
    ws2 = wb.create_sheet("英文条款详情")
    headers2 = ["文件名", "客户姓名", "文件类型", "条款编号", "关键词", "结果", "备注"]
    ws2.append(headers2)
    style_header(ws2)

    # ===== Sheet 3: 中文条款详情 =====
    ws3 = wb.create_sheet("中文条款详情")
    headers3 = ["文件名", "客户姓名", "文件类型", "条款编号", "关键词", "结果", "备注"]
    ws3.append(headers3)
    style_header(ws3)

    # 读取所有条款审核记录（兼容旧的扁平 clause_audit 与新的嵌套 *_audit 格式）
    for f in sorted(RECORDS_DIR.glob("*.json"), reverse=True)[:500]:
        try:
            with open(f, "r", encoding="utf-8") as fp:
                d = json.load(fp)

            # 统一取出条款审核数据：旧格式直接就是 d，新格式嵌套在 clause_audit 里
            if d.get("type") == "clause_audit":
                cd = d
            elif isinstance(d.get("clause_audit"), dict):
                cd = d["clause_audit"]
            else:
                continue

            pass_fill = GREEN if cd.get("overall") == "pass" else RED

            # 汇总行
            row1 = [
                cd.get("file_name", ""),
                cd.get("audit_time", ""),
                cd.get("doc_type", ""),
                "HK" if "HK" in cd.get("doc_type","") else ("US" if "US" in cd.get("doc_type","") else cd.get("doc_type","")),
                cd.get("language", ""),
                cd.get("transaction_type", ""),
                cd.get("customer_name", ""),
                cd.get("account_number", ""),
                cd.get("contract_date", ""),
                cd.get("en_total", 0),
                cd.get("en_passed", 0),
                cd.get("en_total", 0) - cd.get("en_passed", 0),
                cd.get("zh_total", 0),
                cd.get("zh_passed", 0),
                cd.get("zh_total", 0) - cd.get("zh_passed", 0),
                "通过" if cd.get("overall") == "pass" else "失败",
                "; ".join(cd.get("issues", [])),
            ]
            ws1.append(row1)
            style_row(ws1, ws1.max_row, pass_fill)

            # 英文条款行
            for c in cd.get("en_clauses", []):
                r_fill = GREEN if c.get("result") == "pass" else RED
                ws2.append([
                    cd.get("file_name",""),
                    cd.get("customer_name",""),
                    cd.get("doc_type",""),
                    c.get("clause_num",""),
                    c.get("keyword","")[:80],
                    "通过" if c.get("result") == "pass" else "失败",
                    c.get("note",""),
                ])
                style_row(ws2, ws2.max_row, r_fill)

            # 中文条款行
            for c in cd.get("zh_clauses", []):
                r_fill = GREEN if c.get("result") == "pass" else RED
                ws3.append([
                    cd.get("file_name",""),
                    cd.get("customer_name",""),
                    cd.get("doc_type",""),
                    c.get("clause_num",""),
                    c.get("keyword","")[:80],
                    "通过" if c.get("result") == "pass" else "失败",
                    c.get("note",""),
                ])
                style_row(ws3, ws3.max_row, r_fill)

        except Exception:
            continue

    # 调整列宽
    col_widths = {
        ws1: [30,18,10,6,6,8,18,14,12,8,8,8,8,8,8,8,40],
        ws2: [30,18,10,10,60,6,40],
        ws3: [30,18,10,10,60,6,40],
    }
    for ws, widths in col_widths.items():
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clause_audit_report.xlsx"}
    )


# ===== 条款模板管理 =====
CLAUSES_JSON = Path("data/clauses.json")

def _build_clauses_from_constants():
    """从硬编码常量构建条款数据"""
    def to_list(clauses):
        return [[name, kw] for name, kw in clauses]
    empty_langs = {"en": [], "zh_hk": [], "zh_cn": []}
    return {
        "Stock": {
            "order": {"en": to_list(STOCK_EN_CLAUSES), "zh_hk": to_list(STOCK_ZH_HK_CLAUSES), "zh_cn": to_list(STOCK_ZH_CN_CLAUSES)},
            "orderHk": {**empty_langs},
            "companyAction": {**empty_langs},
            "companyActionHk": {**empty_langs},
            "business": {**empty_langs},
            "businessHk": {**empty_langs},
        },
        "Fund":  {"zh_hk": to_list(FUND_ZH_HK_CLAUSES),  "zh_cn": to_list(FUND_ZH_CN_CLAUSES)},
        "VA": {
            "buy":  {"zh_hk": to_list(VA_ZH_HK_CLAUSES), "zh_cn": to_list(VA_ZH_CN_CLAUSES)},
            "sale": {"zh_hk": [], "zh_cn": []},
        },
        "Statement": {
            "monthly": {"zh_hk": [], "zh_cn": []},
            "bbInvest": {"zh_hk": [], "zh_cn": []},
            "southBound": {"zh_hk": [], "zh_cn": []},
        },
    }


@app.get("/api/clauses", response_class=JSONResponse)
async def get_clauses():
    """获取当前条款数据"""
    if CLAUSES_JSON.exists():
        try:
            data = json.loads(CLAUSES_JSON.read_text(encoding="utf-8"))
            return JSONResponse(data)
        except Exception:
            pass
    return JSONResponse(_build_clauses_from_constants())


@app.put("/api/clauses", response_class=JSONResponse)
async def save_clauses(request: Request):
    """保存条款数据"""
    data = await request.json()
    # 验证结构
    # Stock: 嵌套子类型
    if "Stock" not in data:
        return JSONResponse({"error": "缺少产品: Stock"}, status_code=400)
    for st in STOCK_SUBTYPES:
        if st not in data["Stock"]:
            return JSONResponse({"error": f"Stock 缺少子类型: {st}"}, status_code=400)
        for lang in ["en", "zh_hk", "zh_cn"]:
            if lang not in data["Stock"][st]:
                return JSONResponse({"error": f"Stock/{st} 缺少语言: {lang}"}, status_code=400)
            for i, clause in enumerate(data["Stock"][st][lang]):
                if not isinstance(clause, list) or len(clause) != 2:
                    return JSONResponse({"error": f"Stock/{st}/{lang} 条款{i+1}格式错误"}, status_code=400)
    # Fund: 扁平结构（无 en）
    if "Fund" not in data:
        return JSONResponse({"error": "缺少产品: Fund"}, status_code=400)
    for lang in ["zh_hk", "zh_cn"]:
        if lang not in data["Fund"]:
            return JSONResponse({"error": f"Fund 缺少语言: {lang}"}, status_code=400)
        for i, clause in enumerate(data["Fund"][lang]):
            if not isinstance(clause, list) or len(clause) != 2:
                return JSONResponse({"error": f"Fund/{lang} 条款{i+1}格式错误"}, status_code=400)
    # VA: 嵌套子类型（无 en）
    if "VA" not in data:
        return JSONResponse({"error": "缺少产品: VA"}, status_code=400)
    for st in VA_SUBTYPES:
        if st not in data["VA"]:
            return JSONResponse({"error": f"VA 缺少子类型: {st}"}, status_code=400)
        for lang in ["zh_hk", "zh_cn"]:
            if lang not in data["VA"][st]:
                return JSONResponse({"error": f"VA/{st} 缺少语言: {lang}"}, status_code=400)
            for i, clause in enumerate(data["VA"][st][lang]):
                if not isinstance(clause, list) or len(clause) != 2:
                    return JSONResponse({"error": f"VA/{st}/{lang} 条款{i+1}格式错误"}, status_code=400)
    # Statement: 嵌套子类型（无 en，同 VA）
    if "Statement" not in data:
        return JSONResponse({"error": "缺少产品: Statement"}, status_code=400)
    for st in STATEMENT_SUBTYPES:
        if st not in data["Statement"]:
            return JSONResponse({"error": f"Statement 缺少子类型: {st}"}, status_code=400)
        for lang in ["zh_hk", "zh_cn"]:
            if lang not in data["Statement"][st]:
                return JSONResponse({"error": f"Statement/{st} 缺少语言: {lang}"}, status_code=400)
            for i, clause in enumerate(data["Statement"][st][lang]):
                if not isinstance(clause, list) or len(clause) != 2:
                    return JSONResponse({"error": f"Statement/{st}/{lang} 条款{i+1}格式错误"}, status_code=400)
    CLAUSES_JSON.parent.mkdir(parents=True, exist_ok=True)
    CLAUSES_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return JSONResponse({"success": True, "message": "条款已保存"})


@app.post("/api/clauses/import-html", response_class=JSONResponse)
async def import_html_clauses(file: UploadFile = File(...)):
    """从 HTML 模板文件提取所有固定文本作为条款（公司信息、标题、字段标签、重要提示等）"""
    if not file.filename or not file.filename.lower().endswith(('.html', '.htm')):
        return JSONResponse({"error": "仅支持 HTML 文件"}, status_code=400)
    try:
        from bs4 import BeautifulSoup, NavigableString
        import re as _re
        content = await file.read()
        try:
            html_text = content.decode("utf-8")
        except UnicodeDecodeError:
            html_text = content.decode("latin-1")
        soup = BeautifulSoup(html_text, "html.parser")

        clauses = []
        idx = [0]  # mutable counter

        def add(text: str):
            text = text.strip()
            if not text or len(text) < 2:
                return
            # 跳过纯模板变量（如 ${contractNote.xxx}）
            if _re.fullmatch(r'[\s]*(\$\{[^}]+\}\s*)+', text):
                return
            idx[0] += 1
            clauses.append([f"条款{idx[0]}", text])

        # --- 1. 页脚公司信息 (footer_compnay) ---
        footer = soup.find("div", class_="footer_compnay")
        if footer:
            for span in footer.find_all("span"):
                add(span.get_text(strip=True))

        # --- 2. 文档标题 (p.title) ---
        title_p = soup.find("p", class_="title")
        if title_p:
            add(title_p.get_text(strip=True))

        # --- 3. 表头字段标签 (table.table-head) ---
        head_table = soup.find("table", class_="table-head")
        if head_table:
            for td in head_table.find_all("td"):
                raw = td.get_text(strip=True)
                # 提取固定文本标签（去掉模板变量部分）
                # e.g. "Issue Date 發出日期:" from "Issue Date 發出日期:${...}"
                label = _re.sub(r'\$\{[^}]+\}', '', raw).strip().rstrip(':').rstrip('：').strip()
                if label and len(label) >= 4:
                    add(label)

        # --- 4. 主数据表字段标签 (table.table-title td.table-title-tr-title) ---
        title_table = soup.find("table", class_="table-title")
        if title_table:
            for td in title_table.find_all("td", class_="table-title-tr-title"):
                # 获取文本，br 替换为换行符保持可读性
                for br in td.find_all("br"):
                    br.replace_with("\n")
                label = td.get_text(strip=False).strip()
                # 将换行压缩为单个换行
                label = _re.sub(r'\n+', '\n', label).strip()
                if label:
                    add(label)

            # 固定值单元格（如 "Buy/買入", "Sell/賣出"）
            for td in title_table.find_all("td", class_="table-title-tr-content"):
                raw = td.get_text(strip=True)
                # 跳过纯模板变量
                cleaned = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
                if cleaned and not _re.fullmatch(r'[\s]*', cleaned) and len(cleaned) >= 3:
                    add(cleaned)

            # 嵌套 special-table 内的字段标签（部分成交明细）
            special = title_table.find("table", class_="special-table")
            if special:
                for tr in special.find_all("tr", class_="table-special-tr"):
                    tds = tr.find_all("td")
                    if tds:
                        raw = tds[0].get_text(strip=True)
                        # 移除模板变量
                        label = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
                        # 移除尾部括号中的模板占位 e.g. "()"
                        label = _re.sub(r'\(\s*\)\s*$', '', label).strip()
                        if label and len(label) >= 4:
                            add(label)

        # --- 5. 重要提示标题 ---
        tips = soup.find("p", class_="tips-import")
        if tips:
            add(tips.get_text(strip=True))

        # --- 6. 重要提示条款 (div.content with span.pr-8) ---
        content_divs = soup.find_all("div", class_="content")
        for div in content_divs:
            span = div.find("span", class_="pr-8")
            if span:
                num_text = span.get_text(strip=True)
                full_text = div.get_text(strip=True)
                clause_text = full_text[len(num_text):].strip()
                if clause_text and len(clause_text) > 5:
                    add(clause_text)

        if not clauses:
            return JSONResponse({"error": "未能从 HTML 中提取到任何固定文本"}, status_code=400)

        # 从文件名提取版本信息（如 zabank_crypto_contract_note_buy_hk_240910_v2.html -> 2024-09-10 v2）
        version_info = _extract_version_from_filename(file.filename or "")

        return JSONResponse({"success": True, "clauses": clauses, "count": len(clauses), "version": version_info})
    except Exception as e:
        return JSONResponse({"error": f"解析失败: {str(e)}"}, status_code=500)


def _extract_version_from_filename(filename: str) -> str:
    """从文件名提取版本信息，如 zabank_..._240910_v2.html -> 2024-09-10 v2"""
    import re as _re
    name = filename.rsplit('.', 1)[0] if '.' in filename else filename
    # 匹配 YYMMDD_vN 格式
    m = _re.search(r'(\d{6})_v(\d+)', name)
    if m:
        date_str, ver = m.group(1), m.group(2)
        try:
            year = int(date_str[:2]) + 2000
            month = int(date_str[2:4])
            day = int(date_str[4:6])
            return f"{year}-{month:02d}-{day:02d} v{ver}"
        except (ValueError, IndexError):
            pass
    # 匹配 YYYYMMDD_vN 格式
    m = _re.search(r'(\d{8})_v(\d+)', name)
    if m:
        date_str, ver = m.group(1), m.group(2)
        try:
            return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]} v{ver}"
        except (ValueError, IndexError):
            pass
    # 匹配单独的 vN
    m = _re.search(r'_v(\d+)', name)
    if m:
        return f"v{m.group(1)}"
    return filename


def _extract_clauses_from_html(html_text: str) -> list:
    """从 HTML 文本提取条款列表，复用现有逻辑"""
    from bs4 import BeautifulSoup
    import re as _re

    soup = BeautifulSoup(html_text, "html.parser")
    clauses = []
    idx = [0]

    def add(text: str):
        text = text.strip()
        if not text or len(text) < 2:
            return
        if _re.fullmatch(r'[\s]*(\$\{[^}]+\}\s*)+', text):
            return
        idx[0] += 1
        clauses.append([f"条款{idx[0]}", text])

    footer = soup.find("div", class_="footer_compnay")
    if footer:
        for span in footer.find_all("span"):
            add(span.get_text(strip=True))

    title_p = soup.find("p", class_="title")
    if title_p:
        add(title_p.get_text(strip=True))

    head_table = soup.find("table", class_="table-head")
    if head_table:
        for td in head_table.find_all("td"):
            raw = td.get_text(strip=True)
            label = _re.sub(r'\$\{[^}]+\}', '', raw).strip().rstrip(':').rstrip('：').strip()
            if label and len(label) >= 4:
                add(label)

    title_table = soup.find("table", class_="table-title")
    if title_table:
        for td in title_table.find_all("td", class_="table-title-tr-title"):
            for br in td.find_all("br"):
                br.replace_with("\n")
            label = td.get_text(strip=False).strip()
            label = _re.sub(r'\n+', '\n', label).strip()
            if label:
                add(label)
        for td in title_table.find_all("td", class_="table-title-tr-content"):
            raw = td.get_text(strip=True)
            cleaned = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
            if cleaned and not _re.fullmatch(r'[\s]*', cleaned) and len(cleaned) >= 3:
                add(cleaned)
        special = title_table.find("table", class_="special-table")
        if special:
            for tr in special.find_all("tr", class_="table-special-tr"):
                tds = tr.find_all("td")
                if tds:
                    raw = tds[0].get_text(strip=True)
                    label = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
                    label = _re.sub(r'\(\s*\)\s*$', '', label).strip()
                    if label and len(label) >= 4:
                        add(label)

    tips = soup.find("p", class_="tips-import")
    if tips:
        add(tips.get_text(strip=True))

    content_divs = soup.find_all("div", class_="content")
    for div in content_divs:
        span = div.find("span", class_="pr-8")
        if span:
            num_text = span.get_text(strip=True)
            full_text = div.get_text(strip=True)
            clause_text = full_text[len(num_text):].strip()
            if clause_text and len(clause_text) > 5:
                add(clause_text)

    return clauses


def _detect_product_from_path(html_path: Path) -> tuple:
    """
    从文件路径识别 (product, subtype, lang)
    文件夹结构约定:
      Fund Contract note/        -> Fund, None
      investMonthlyStatement/    -> Statement, monthly/bbInvest/southBound
      VA Contract note/          -> VA, buy/sale
      Stock .../order/           -> Stock, order
      Stock .../businessHk/      -> Stock, businessHk
      ... 等
    语言从文件名识别: _hk_ -> zh_hk, _zh_ -> zh_cn
    """
    parts = [p.lower() for p in html_path.parts]
    fname = html_path.name

    lang = _detect_lang_from_filename(fname)

    # Stock 优先：parent 文件夹名是子类型（最明确）
    stock_subtypes = {"order", "orderhk", "companyaction", "companyactionhk", "business", "businesshk"}
    parent = html_path.parent.name.lower()
    if parent in stock_subtypes:
        subtype_map = {
            "order": "order", "orderhk": "orderHk",
            "companyaction": "companyAction", "companyactionhk": "companyActionHk",
            "business": "business", "businesshk": "businessHk",
        }
        return "Stock", subtype_map[parent], lang

    # Fund
    if any("fund" in p for p in parts):
        return "Fund", None, lang

    # Statement / Monthly（仅当直接父文件夹包含 monthly/statement，避免误匹配 Stock 上级目录）
    if any("monthly" in p or "investmonthlystatement" in p for p in parts):
        subtype = _detect_subtype_from_filename(fname)
        return "Statement", subtype, lang

    # VA
    if any("va" in p or "crypto" in p or "virtual" in p for p in parts):
        sub = "sale" if ("sale" in fname.lower()) else "buy"
        return "VA", sub, lang

    return None, None, lang


@app.post("/api/clauses/sync-folder", response_class=JSONResponse)
async def sync_folder(request: Request):
    """扫描本地文件夹，自动识别并导入所有 HTML 模板"""
    body = await request.json()
    folder_path = body.get("folder_path", "").strip()
    if not folder_path:
        return JSONResponse({"error": "未提供文件夹路径"}, status_code=400)

    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        return JSONResponse({"error": f"文件夹不存在: {folder_path}"}, status_code=400)

    # 递归找所有 HTML 文件
    html_files = list(folder.rglob("*.html")) + list(folder.rglob("*.htm"))
    if not html_files:
        return JSONResponse({"error": "文件夹内未找到任何 HTML 文件"}, status_code=400)

    # 加载现有条款数据
    if CLAUSES_JSON.exists():
        try:
            clauses_data = json.loads(CLAUSES_JSON.read_text(encoding="utf-8"))
        except Exception:
            clauses_data = _build_clauses_from_constants()
    else:
        clauses_data = _build_clauses_from_constants()

    results = []
    errors = []

    for html_path in sorted(html_files):
        product, subtype, lang = _detect_product_from_path(html_path)
        if not product or not lang:
            errors.append({"file": html_path.name, "error": f"无法识别产品或语言 (product={product}, lang={lang})"})
            continue

        try:
            try:
                html_text = html_path.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                html_text = html_path.read_text(encoding="latin-1")

            clauses = _extract_clauses_from_html(html_text)
            if not clauses:
                errors.append({"file": html_path.name, "error": "未能提取到任何条款"})
                continue

            version = _extract_version_from_filename(html_path.name)

            # 写入 clauses_data
            if product == "Stock":
                if "Stock" not in clauses_data:
                    clauses_data["Stock"] = {}
                if subtype not in clauses_data["Stock"]:
                    clauses_data["Stock"][subtype] = {"en": [], "zh_hk": [], "zh_cn": []}
                clauses_data["Stock"][subtype][lang] = clauses
                if version:
                    clauses_data["Stock"][subtype].setdefault("_meta", {})[lang] = {"version": version}
            elif product in ("Fund", "Statement") and subtype is None:
                # Fund: flat
                if product not in clauses_data:
                    clauses_data[product] = {}
                clauses_data[product][lang] = clauses
                if version:
                    clauses_data[product].setdefault("_meta", {})[lang] = {"version": version}
            else:
                # VA / Statement with subtype
                if product not in clauses_data:
                    clauses_data[product] = {}
                if subtype not in clauses_data[product]:
                    clauses_data[product][subtype] = {}
                clauses_data[product][subtype][lang] = clauses
                if version:
                    clauses_data[product][subtype].setdefault("_meta", {})[lang] = {"version": version}

            results.append({
                "file": html_path.name,
                "product": product,
                "subtype": subtype,
                "lang": lang,
                "count": len(clauses),
                "version": version,
            })
        except Exception as e:
            errors.append({"file": html_path.name, "error": str(e)})

    # 保存
    if results:
        CLAUSES_JSON.parent.mkdir(parents=True, exist_ok=True)
        CLAUSES_JSON.write_text(json.dumps(clauses_data, ensure_ascii=False, indent=2), encoding="utf-8")

    return JSONResponse({
        "success": True,
        "total": len(html_files),
        "imported": len(results),
        "errors_count": len(errors),
        "results": results,
        "errors": errors,
        "summary": f"共扫描 {len(html_files)} 个文件，成功导入 {len(results)} 个" + (f"，{len(errors)} 个失败" if errors else ""),
    })


def _detect_lang_from_filename(filename: str) -> str:
    """从文件名识别语言: _hk_ -> zh_hk, _zh_ -> zh_cn"""
    name = filename.lower()
    if "_hk_" in name or "_hk." in name:
        return "zh_hk"
    if "_zh_" in name or "_zh." in name:
        return "zh_cn"
    return ""


def _detect_subtype_from_filename(filename: str) -> str:
    """从文件名识别子类型"""
    name = filename.lower()
    if "bbinvest" in name:
        return "bbInvest"
    if "southbound" in name:
        return "southBound"
    return "monthly"


@app.post("/api/clauses/import-html-batch", response_class=JSONResponse)
async def import_html_clauses_batch(files: list[UploadFile] = File(...)):
    """批量导入 HTML 模板文件，自动识别语言和子类型"""
    if not files:
        return JSONResponse({"error": "未选择文件"}, status_code=400)

    from bs4 import BeautifulSoup, NavigableString
    import re as _re

    results = []
    errors = []

    for file in files:
        if not file.filename or not file.filename.lower().endswith(('.html', '.htm')):
            errors.append({"file": file.filename, "error": "非 HTML 文件"})
            continue

        try:
            content = await file.read()
            try:
                html_text = content.decode("utf-8")
            except UnicodeDecodeError:
                html_text = content.decode("latin-1")
            soup = BeautifulSoup(html_text, "html.parser")

            clauses = []
            idx = [0]

            def add(text: str):
                text = text.strip()
                if not text or len(text) < 2:
                    return
                if _re.fullmatch(r'[\s]*(\$\{[^}]+\}\s*)+', text):
                    return
                idx[0] += 1
                clauses.append([f"条款{idx[0]}", text])

            # 复用现有提取逻辑（6个提取区域）
            footer = soup.find("div", class_="footer_compnay")
            if footer:
                for span in footer.find_all("span"):
                    add(span.get_text(strip=True))

            title_p = soup.find("p", class_="title")
            if title_p:
                add(title_p.get_text(strip=True))

            head_table = soup.find("table", class_="table-head")
            if head_table:
                for td in head_table.find_all("td"):
                    raw = td.get_text(strip=True)
                    label = _re.sub(r'\$\{[^}]+\}', '', raw).strip().rstrip(':').rstrip('：').strip()
                    if label and len(label) >= 4:
                        add(label)

            title_table = soup.find("table", class_="table-title")
            if title_table:
                for td in title_table.find_all("td", class_="table-title-tr-title"):
                    for br in td.find_all("br"):
                        br.replace_with("\n")
                    label = td.get_text(strip=False).strip()
                    label = _re.sub(r'\n+', '\n', label).strip()
                    if label:
                        add(label)
                for td in title_table.find_all("td", class_="table-title-tr-content"):
                    raw = td.get_text(strip=True)
                    cleaned = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
                    if cleaned and not _re.fullmatch(r'[\s]*', cleaned) and len(cleaned) >= 3:
                        add(cleaned)
                special = title_table.find("table", class_="special-table")
                if special:
                    for tr in special.find_all("tr", class_="table-special-tr"):
                        tds = tr.find_all("td")
                        if tds:
                            raw = tds[0].get_text(strip=True)
                            label = _re.sub(r'\$\{[^}]+\}', '', raw).strip()
                            label = _re.sub(r'\(\s*\)\s*$', '', label).strip()
                            if label and len(label) >= 4:
                                add(label)

            tips = soup.find("p", class_="tips-import")
            if tips:
                add(tips.get_text(strip=True))

            content_divs = soup.find_all("div", class_="content")
            for div in content_divs:
                span = div.find("span", class_="pr-8")
                if span:
                    num_text = span.get_text(strip=True)
                    full_text = div.get_text(strip=True)
                    clause_text = full_text[len(num_text):].strip()
                    if clause_text and len(clause_text) > 5:
                        add(clause_text)

            if not clauses:
                errors.append({"file": file.filename, "error": "未能提取到任何条款"})
                continue

            lang = _detect_lang_from_filename(file.filename)
            subtype = _detect_subtype_from_filename(file.filename)
            version = _extract_version_from_filename(file.filename)

            results.append({
                "filename": file.filename,
                "subtype": subtype,
                "lang": lang,
                "clauses": clauses,
                "count": len(clauses),
                "version": version,
            })
        except Exception as e:
            errors.append({"file": file.filename, "error": str(e)})

    if not results and errors:
        return JSONResponse({"error": f"所有文件处理失败", "errors": errors}, status_code=400)

    return JSONResponse({
        "success": True,
        "results": results,
        "errors": errors,
        "summary": f"成功处理 {len(results)} 个文件" + (f"，{len(errors)} 个失败" if errors else ""),
    })


# ===== 产品审核（统一端点） =====
@app.post("/api/product-audit", response_class=JSONResponse)
async def product_audit(files: list[UploadFile] = File(...), product: str = Form(...), subtype: Optional[str] = Form(None), skip_ai: bool = Form(False)):
    """统一产品审核端点：条款审核 + AI/VA 审核。product=auto 时自动识别每个文件的类型"""
    if product not in ("stock", "fund", "va", "statement", "auto"):
        return JSONResponse({"error": f"不支持的产品类型: {product}"}, status_code=400)
    if product == "stock" and subtype and subtype not in STOCK_SUBTYPES:
        return JSONResponse({"error": f"不支持的股票子类型: {subtype}"}, status_code=400)
    if product == "va" and subtype and subtype not in VA_SUBTYPES:
        return JSONResponse({"error": f"不支持的VA子类型: {subtype}"}, status_code=400)
    if product == "statement" and subtype and subtype not in STATEMENT_SUBTYPES:
        return JSONResponse({"error": f"不支持的月结单子类型: {subtype}"}, status_code=400)
    if product not in ("stock", "va", "statement") and subtype:
        return JSONResponse({"error": "仅股票、VA和月结单产品支持子类型参数"}, status_code=400)
    if len(files) > 200:
        return JSONResponse({"error": f"单次最多上传 200 份文件，当前 {len(files)} 份"}, status_code=400)

    auditor = ContractAuditor()

    # 并行审核：所有文件并发处理，避免逐个串行等待 AI 调用
    def process_one(file) -> dict:
        if not file.filename or not file.filename.lower().endswith('.pdf'):
            return {"error": {"file": file.filename, "error": "非 PDF 文件"}}

        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        safe_name = Path(file.filename).name or "file.pdf"
        file_path = UPLOAD_DIR / f"{ts}_{safe_name}"

        # 文件字节需要在 async 上下文中读取，这里读的是 UploadFile 的同步副本
        content = file._read_bytes if hasattr(file, "_read_bytes") else None
        try:
            with open(file_path, "wb") as f:
                f.write(content)

            item = {"file_name": file.filename.replace("\\", "/")}

            # 1. 条款审核（auto 模式下不传 subtype，由 auditor 自动识别）
            audit_subtype = None if product == "auto" else subtype
            clause_result = auditor.audit(str(file_path), file.filename, subtype=audit_subtype)
            item["clause_audit"] = clause_result.model_dump()

            # auto 模式：根据识别的 doc_type 映射到实际产品
            effective_product = product
            if product == "auto":
                dt = clause_result.doc_type or ""
                if dt == "VA":
                    effective_product = "va"
                elif dt == "Fund":
                    effective_product = "fund"
                elif dt == "Statement":
                    effective_product = "statement"
                elif dt in ("Stock-HK", "Stock-US", "Unknown"):
                    effective_product = "stock"
                else:
                    effective_product = "stock"
            item["product"] = effective_product

            # 2. AI 审核（股票 + 基金）
            if effective_product in ("stock", "fund") and not skip_ai:
                try:
                    parser = PDFParser()
                    contract_data = parser.parse(str(file_path))
                    engine = get_audit_engine()
                    report = engine.audit(contract_data, file.filename)
                    item["ai_audit"] = {
                        "findings": [f.model_dump() for f in report.findings],
                        "total_findings": report.total_findings,
                        "critical_count": report.critical_count,
                        "warning_count": report.warning_count,
                        "info_count": report.info_count,
                        "ai_summary": report.ai_summary,
                        "fee_waiver_details": report.fee_waiver_details,
                    }
                except Exception:
                    item["ai_audit"] = None

            # 3. VA 专项审核
            if effective_product == "va":
                try:
                    checker = VAChecker()
                    va_result = checker.check(str(file_path), file.filename)
                    item["va_audit"] = va_result.model_dump()
                except Exception:
                    item["va_audit"] = None

            save_record({"type": f"{effective_product}_audit", **item})
            return {"result": item}
        except Exception as e:
            return {"error": {"file": file.filename, "error": str(e)}}

    # 读取所有文件字节（async，仅读取）
    for file in files:
        file._read_bytes = await file.read()

    # 并发执行，IO 密集型适当放宽线程数
    workers = min(len(files), 10)
    loop = asyncio.get_running_loop()
    with ThreadPoolExecutor(max_workers=workers) as pool:
        outcomes = await asyncio.gather(*[loop.run_in_executor(pool, process_one, f) for f in files])

    results = []
    errors = []
    for outcome in outcomes:
        if "result" in outcome:
            results.append(outcome["result"])
        elif "error" in outcome:
            errors.append(outcome["error"])

    return {"success": True, "total": len(results), "errors": errors, "results": results}


@app.get("/api/export/excel", response_class=StreamingResponse)
async def export_excel():
    """导出所有审核记录为 Excel"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()

    # ===== Sheet 1: 成交单审核汇总 =====
    ws1 = wb.active
    ws1.title = "成交单审核"
    headers1 = ["文件名", "审核时间", "客户姓名", "账户号码", "股票", "成交日期",
                "严重问题", "警告", "信息", "总计", "AI摘要"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # ===== Sheet 2: VA审核汇总 =====
    ws2 = wb.create_sheet("VA成交单审核")
    headers2 = ["文件名", "审核时间", "客户姓名", "账户号码", "资产名称", "资产代码",
                "语言", "交易类型", "成交日期", "总检查项", "通过", "失败", "总体结论"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center")

    # ===== Sheet 3: VA详细检查项 =====
    ws3 = wb.create_sheet("VA详细检查项")
    headers3 = ["文件名", "客户姓名", "检查类别", "检查项目", "结果", "详情"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center")

    # 读取所有记录
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")

    for f in sorted(RECORDS_DIR.glob("*.json"), reverse=True)[:200]:
        try:
            with open(f, "r", encoding="utf-8") as fp:
                data = json.load(fp)

            # VA 记录：旧格式 flat va，新格式嵌套 va_audit
            va = None
            if data.get("type") == "va":
                va = data
            elif isinstance(data.get("va_audit"), dict):
                va = data["va_audit"]

            if va:
                row = [
                    va.get("file_name", ""),
                    va.get("audit_time", ""),
                    va.get("customer_name", ""),
                    va.get("account_number", ""),
                    va.get("asset_name", ""),
                    va.get("asset_code", ""),
                    va.get("language", ""),
                    va.get("transaction_type", ""),
                    va.get("contract_date", ""),
                    va.get("total", 0),
                    va.get("passed", 0),
                    va.get("failed", 0),
                    "通过" if va.get("overall") == "pass" else "失败",
                ]
                r = ws2.append(row)
                last = ws2.max_row
                fill = pass_fill if va.get("overall") == "pass" else fail_fill
                for cell in ws2[last]:
                    cell.fill = fill

                # 写详细检查项
                for check in va.get("checks", []):
                    ws3.append([
                        va.get("file_name", ""),
                        va.get("customer_name", ""),
                        check.get("category", ""),
                        check.get("item", ""),
                        check.get("result", ""),
                        check.get("detail", ""),
                    ])
                    last3 = ws3.max_row
                    r = check.get("result", "")
                    f3 = pass_fill if r == "pass" else (fail_fill if r == "fail" else warn_fill)
                    ws3[last3][4].fill = f3

            # 成交单 AI 审核记录：旧格式 success+contract_data，新格式嵌套 ai_audit
            ai = None
            if data.get("success") and isinstance(data.get("contract_data"), dict):
                ai = {"data": data, "cd": data.get("contract_data", {})}
            elif isinstance(data.get("ai_audit"), dict):
                ai = {"data": data.get("ai_audit", {}), "cd": data.get("clause_audit", {}) or {}}

            if ai:
                d_ai = ai["data"]
                cd = ai["cd"]
                row = [
                    data.get("file_name", cd.get("file_name", "")),
                    cd.get("audit_time", data.get("audit_time", "")),
                    f"{cd.get('customer_name','')}".strip(),
                    cd.get("account_number", ""),
                    "",
                    cd.get("contract_date", ""),
                    d_ai.get("critical_count", 0),
                    d_ai.get("warning_count", 0),
                    d_ai.get("info_count", 0),
                    d_ai.get("total_findings", 0),
                    d_ai.get("ai_summary", "")[:100],
                ]
                ws1.append(row)
                last = ws1.max_row
                fill = pass_fill if d_ai.get("critical_count", 0) == 0 else fail_fill
                for cell in ws1[last]:
                    cell.fill = fill
        except Exception:
            continue

    # 调整列宽
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_report.xlsx"}
    )


@app.get("/api/health")
async def health_check():
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    return {
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "ai_enabled": bool(api_key),
        "ai_provider": "OpenRouter Nemotron" if api_key else "未配置",
    }


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
